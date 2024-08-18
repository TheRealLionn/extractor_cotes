import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import pandas as pd
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Variables globales pour stocker les chemins des fichiers et les colonnes
file_path1 = ""
file_path2 = ""
cote_column = ""
target_column = ""

# Couleurs et styles inspirés de Facebook
bg_color = "#f0f2f5"
button_color = "#4267B2"
button_text_color = "#ffffff"
font_family = "Helvetica"


# Fonction pour charger le premier fichier
def load_file1():
    global file_path1
    file_path1 = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx"), ("PDF files", "*.pdf"), ("Image files", "*.png;*.jpg;*.jpeg")])
    if file_path1:
        label_file1.config(text=f"Fichier 1 : {os.path.basename(file_path1)}")
        display_column_titles(file_path1, treeview_columns1)
    else:
        messagebox.showwarning("Avertissement", "Aucun fichier sélectionné")


# Fonction pour charger le deuxième fichier
def load_file2():
    global file_path2
    file_path2 = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path2:
        label_file2.config(text=f"Fichier 2 : {os.path.basename(file_path2)}")
        display_column_titles(file_path2, treeview_columns2)
    else:
        messagebox.showwarning("Avertissement", "Aucun fichier sélectionné")


# Fonction pour vider les fichiers chargés
def clear_files():
    global file_path1, file_path2, cote_column, target_column
    file_path1 = ""
    file_path2 = ""
    cote_column = ""
    target_column = ""
    label_file1.config(text="Fichier 1 : Aucun fichier sélectionné")
    label_file2.config(text="Fichier 2 : Aucun fichier sélectionné")
    label_cote_column.config(text="")
    label_target_column.config(text="")
    treeview.delete(*treeview.get_children())
    treeview_columns1.delete(*treeview_columns1.get_children())
    treeview_columns2.delete(*treeview_columns2.get_children())
    messagebox.showinfo("Info", "Fichiers chargés vidés")


# Fonction pour traiter les fichiers
def process_files():
    global cote_column, target_column
    if not file_path1 or not file_path2:
        messagebox.showwarning("Avertissement", "Veuillez sélectionner les deux fichiers")
        return

    if not cote_column:
        messagebox.showwarning("Avertissement", "Veuillez sélectionner la colonne des cotes dans le premier fichier")
        return

    if not target_column:
        messagebox.showwarning("Avertissement", "Veuillez sélectionner la colonne cible dans le deuxième fichier")
        return

    extension = os.path.splitext(file_path1)[1].lower()
    if extension in ['.xlsx']:
        df1 = pd.read_excel(file_path1)
    elif extension in ['.pdf']:
        df1 = read_pdf(file_path1)
    elif extension in ['.png', '.jpg', '.jpeg']:
        df1 = read_image(file_path1)
    else:
        messagebox.showerror("Erreur", "Type de fichier non supporté")
        return

    if 'matricule' not in df1.columns:
        messagebox.showerror("Erreur", "Le premier fichier doit contenir une colonne 'matricule'")
        return

    df2 = pd.read_excel(file_path2)
    if 'matricule' not in df2.columns:
        messagebox.showerror("Erreur", "Le deuxième fichier doit contenir une colonne 'matricule'")
        return

    # Vérification des matricules manquants
    missing_in_df1 = set(df2['matricule']) - set(df1['matricule'])
    missing_in_df2 = set(df1['matricule']) - set(df2['matricule'])

    if missing_in_df1 or missing_in_df2:
        confirmation_message = "Il y a des matricules qui ne correspondent pas entre les deux fichiers.\n"
        if missing_in_df1:
            confirmation_message += f"Matricules dans le deuxième fichier mais pas dans le premier: {', '.join(map(str, missing_in_df1))}\n"
        if missing_in_df2:
            confirmation_message += f"Matricules dans le premier fichier mais pas dans le deuxième: {', '.join(map(str, missing_in_df2))}\n"
        confirmation_message += "Voulez-vous continuer le traitement?"

        if not messagebox.askyesno("Confirmation", confirmation_message):
            return

    missing_matricules = update_points(df1, df2)
    save_updated_file(df1, df2, missing_in_df2)

    if missing_matricules:
        show_missing_matricules(missing_matricules)
        messagebox.showinfo("Succès",
                            "Le fichier a été mis à jour avec succès, mais certains matricules étaient manquants.")
    else:
        messagebox.showinfo("Succès", "Le fichier a été mis à jour avec succès.")


# Fonction pour lire le PDF
def read_pdf(file_path):
    document = fitz.open(file_path)
    text = ""
    for page_num in range(len(document)):
        page = document.load_page(page_num)
        text += page.get_text()
    return parse_text_to_df(text)


# Fonction pour lire l'image
def read_image(file_path):
    image = Image.open(file_path)
    text = pytesseract.image_to_string(image)
    return parse_text_to_df(text)


# Fonction pour parser le texte en DataFrame
def parse_text_to_df(text):
    lines = text.split('\n')
    data = []
    for line in lines:
        if line.strip():
            parts = line.split()
            if len(parts) >= 3:
                matricule, nom, cote = parts[0], ' '.join(parts[1:-1]), parts[-1]
                data.append([matricule, nom, cote])
    return pd.DataFrame(data, columns=['matricule', 'nom', 'cotes'])


# Fonction pour mettre à jour les points dans le deuxième fichier
def update_points(df1, df2):
    df1.set_index('matricule', inplace=True)
    missing_matricules = []

    def update_row(row):
        if row['matricule'] in df1.index:
            return df1.at[row['matricule'], cote_column]
        else:
            missing_matricules.append((row['matricule'], row[target_column]))
            return row[target_column]

    df2[target_column] = df2.apply(update_row, axis=1)
    return missing_matricules


# Fonction pour sauvegarder le fichier Excel mis à jour et ajouter les nouveaux matricules en rouge
def save_updated_file(df1, df2, missing_in_df2):
    with pd.ExcelWriter(file_path2, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df2.to_excel(writer, index=False, sheet_name='Sheet1')

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Ajouter les matricules manquants à la fin et les colorer en rouge
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        for matricule in missing_in_df2:
            new_row = df1.loc[matricule].copy()
            new_row[target_column] = df1.at[matricule, cote_column]
            row_index = worksheet.max_row + 1
            for col_index, value in enumerate(new_row, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=value)
                cell.fill = red_fill


# Fonction pour afficher les matricules manquants dans le tableau
def show_missing_matricules(missing_matricules):
    treeview.delete(*treeview.get_children())  # Clear existing items
    for matricule, old_value in missing_matricules:
        treeview.insert("", "end", values=(matricule, old_value))


# Fonction pour afficher les titres des colonnes dans un tableau
def display_column_titles(file_path, treeview_columns):
    treeview_columns.delete(*treeview_columns.get_children())
    extension = os.path.splitext(file_path)[1].lower()

    if extension in ['.xlsx']:
        df = pd.read_excel(file_path)
    elif extension in ['.pdf']:
        df = read_pdf(file_path)
    elif extension in ['.png', '.jpg', '.jpeg']:
        df = read_image(file_path)
    else:
        messagebox.showerror("Erreur", "Type de fichier non supporté")
        return

    columns = df.columns.tolist()
    treeview_columns["columns"] = columns
    for col in columns:
        treeview_columns.heading(col, text=col)
        treeview_columns.column(col, width=100, anchor="center")
    treeview_columns.insert("", "end", values=columns)


# Fonction pour gérer la sélection des colonnes
def on_column_click(event, treeview, file_number):
    global cote_column, target_column
    selected_item = treeview.identify_column(event.x)
    column_index = int(selected_item.replace("#", "")) - 1
    column_name = treeview["columns"][column_index]

    if file_number == 1:
        cote_column = column_name
        label_cote_column.config(text=f"Colonne sélectionnée : {cote_column}")
    elif file_number == 2:
        target_column = column_name
        label_target_column.config(text=f"Colonne sélectionnée : {target_column}")


# Interface graphique
root = tk.Tk()
root.title("Application de Récupération des Cotes")
root.config(bg=bg_color)

# Configuration de la grille pour la responsivité
root.grid_columnconfigure(0, weight=1)
root.grid_rowconfigure(1, weight=1)
root.grid_rowconfigure(3, weight=1)

frame = tk.Frame(root, bg=bg_color)
frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

label_instruction = tk.Label(frame, text="Sélectionnez les fichiers pour récupérer et mettre à jour les cotes",
                             bg=bg_color, font=(font_family, 12))
label_instruction.pack(pady=10)

button_file1 = tk.Button(frame, text="Charger le premier fichier", command=load_file1, bg=button_color,
                         fg=button_text_color, font=(font_family, 10))
button_file1.pack(pady=10, fill="x")

label_file1 = tk.Label(frame, text="Fichier 1 : Aucun fichier sélectionné", bg=bg_color, font=(font_family, 10))
label_file1.pack(pady=10, fill="x")

label_cote_column = tk.Label(frame, text="", bg=bg_color, font=(font_family, 10))
label_cote_column.pack(pady=10, fill="x")

# Ajout du tableau pour afficher les titres des colonnes du premier fichier
treeview_columns1 = ttk.Treeview(frame, columns=(), show="headings", height=1)
treeview_columns1.pack(pady=10, fill="x")
treeview_columns1.bind("<ButtonRelease-1>", lambda event: on_column_click(event, treeview_columns1, 1))

button_file2 = tk.Button(frame, text="Charger le deuxième fichier", command=load_file2, bg=button_color,
                         fg=button_text_color, font=(font_family, 10))
button_file2.pack(pady=10, fill="x")

label_file2 = tk.Label(frame, text="Fichier 2 : Aucun fichier sélectionné", bg=bg_color, font=(font_family, 10))
label_file2.pack(pady=10, fill="x")

label_target_column = tk.Label(frame, text="", bg=bg_color, font=(font_family, 10))
label_target_column.pack(pady=10, fill="x")

# Ajout du tableau pour afficher les titres des colonnes du deuxième fichier
treeview_columns2 = ttk.Treeview(frame, columns=(), show="headings", height=1)
treeview_columns2.pack(pady=10, fill="x")
treeview_columns2.bind("<ButtonRelease-1>", lambda event: on_column_click(event, treeview_columns2, 2))

button_process = tk.Button(frame, text="Traiter les fichiers", command=process_files, bg=button_color,
                           fg=button_text_color, font=(font_family, 10))
button_process.pack(pady=10, fill="x")

button_clear = tk.Button(frame, text="Vider les fichiers chargés", command=clear_files, bg=button_color,
                         fg=button_text_color, font=(font_family, 10))
button_clear.pack(pady=10, fill="x")

# Section pour le tableau affichant les matricules manquants
label_missing = tk.Label(root, text="Matricules non trouvés et données non copiées :", bg=bg_color,
                         font=(font_family, 12))
label_missing.grid(row=2, column=0, padx=10, pady=10, sticky="ew")

# Ajout d'une barre de défilement pour le tableau des matricules manquants
frame_missing = tk.Frame(root, bg=bg_color)
frame_missing.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")

scrollbar = tk.Scrollbar(frame_missing)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

treeview = ttk.Treeview(frame_missing, columns=("Matricule", "Valeur Ancienne"), show="headings", height=10,
                        yscrollcommand=scrollbar.set)
treeview.heading("Matricule", text="Matricule")
treeview.heading("Valeur Ancienne", text="Valeur Ancienne")
treeview.column("Matricule", width=150)
treeview.column("Valeur Ancienne", width=150)
treeview.pack(fill="both", expand=True)

scrollbar.config(command=treeview.yview)

# Style de la table
style = ttk.Style()
style.configure("Treeview", font=(font_family, 10), rowheight=25, background=bg_color, fieldbackground=bg_color)
style.configure("Treeview.Heading", font=(font_family, 11, "bold"), background=button_color,
                foreground=button_text_color)

root.mainloop()
